import logging
import os
import subprocess
import glob
from openpyxl import load_workbook
import argparse
import logging
import os

import coloredlogs

def setLogger(mode, module):
    """Define the log that we will use."""

    # create logger with 'spam_application'
    logger = logging.getLogger(module)
    logger.setLevel(logging.DEBUG)

    log = logging.StreamHandler()

    # Select the mode
    if mode is 'Debug':
        coloredlogs.install(level='DEBUG')
    elif mode is 'Production':
        coloredlogs.install()

    formatter = logging.Formatter('#[%(levelname)s] - %(name)s - %(message)s')
    log.setFormatter(formatter)

    # logger.addHandler(log)

    # read database here
    logger.info("Entering " + mode + " mode")

    return(logger)


def setMode(args):
    if args.d is True:
        mode = 'Debug'
    else:
        mode = 'Production'

    return(mode)


def check_exist(input_path, logger):
    if not os.path.exists(input_path):
        logger.error('File "' + input_path + '" does not exists')
        logger.error('Exiting...')
        exit(1)

    return(None)


def check_already_exist(input_path, logger):
    if os.path.exists(input_path):
        logger.error('File "' + input_path + '" already exists, cannot overwrite it')
        logger.error('Exiting...')
        exit(1)

    return(None)


def create_folders(dir_list, logger):
    for folder in dir_list:
        if not os.path.exists(folder):
            cmd = 'mkdir ' + folder
            logger.debug(cmd)
            run_cmd(cmd, 1)


def check_input(starting_files, logger):
    if type(starting_files) == str:
        if not os.path.exists(starting_files):
            logger.error('Input file does not exists -> ' + starting_files)
            exit(1)

    elif type(starting_files) == list:
        for in_file in starting_files:
            if not os.path.exists(in_file):
                logger.error('Input file does not exists -> ' + in_file)
                exit(1)


def get_build_for_reference_genome(args):
    """
    Function that checks the arguments and retrieves the build of the
    reference genome to use for the analysis.

    If none is provided, use hs37
    """
    if args.build == None:
        return('GRCh37')
    else:
        if args.build in ['GRCh38', 'GRCh37']:
            return(args.build)
        else:
            logger.error('Build ' + args.build + ' not found')
            logger.error('Exiting...')
            exit(1)


def get_aditional_config(args, config, logger):
    config['conf'] = {}
    config['conf']['build'] = get_build_for_reference_genome(args)

    return(config)


def joinPath(pathItems):
    path = os.path.join(*pathItems)
    return(os.path.normpath(path))


def get_samples_from_registry(args, registry, config, modality, logger, cbio):

    samples, assayID = excel_qc(registry, config, logger, cbio)

    return(samples, assayID)



def recover_fastq_files(fastq_folder, sampleID, config, logger):
    logger.debug("Looking for FASTQ files in directory -> " + fastq_folder)

    if not os.path.exists(fastq_folder):
        logger.warn('Input fastq folder does not exists -> ' + fastq_folder)
        logger.warn('Exiting...')
        exit(1)

    input_fastq_files = glob.glob(os.path.join(fastq_folder, sampleID + '*'))

    if len(input_fastq_files) < 2:
        logger.warn('List of fastq files recovered has fewer than 2 files -> ' + str(input_fastq_files))
        logger.warn('Exiting...')
        exit(1)

    return(input_fastq_files)



##################################################
### TODO PUT INTO FILE AND REMOVE DEPENDENCIES ###
##################################################

def col_names(ws, logger):
    """
    Function to check if column names are correct
    Column names are located at row 9, columns CDEFGHIJKLMNO
    """
    col_names = []
    ref_cols = ['Muestra',
                'Genes',
                'Cód Plásmido',
                'Nº Amplicones',
                'Proporción (%)',
                'Qubit (ng/ul)',
                'Vf (µl)',
                'Cf (ng/uL)',
                'Vi (µl)',
                'Agua (ul)',
                'Index 1',
                'Index 2',
                'Determinación']

    for column in "CDEFGHIJKLMNO":
        cell = "{}{}".format(column, 9)
        col_names.append(ws[cell].value)

    if col_names != ref_cols:
        logger.error('Los nombres de las columnas del PNT no son los esperados')
        logger.error('Esperamos que las columnas sean: ')
        logger.error(ref_cols)
        logger.error('El usuario ha introducido: ')
        logger.error(col_names)
        logger.error('Saliendo del programa...')
        exit(1)

    else:
        logger.info('Success reading excel columns from registry')


def pnt_version(ws, logger):
    """
    Function to check if PNT has the expected version
    PNT version is located at cell E2 on the excel input file
    """
    version = str(ws['E2'].value)
    if "R-PNT-TEC-016-1/rev.3" not in version:
        logger.error('La celda E2 del PNT no contiene la version R-PNT-TEC-016-1/rev.3')
        logger.error('La version del documento excel introducido es: ')
        logger.error(version)
        logger.error('Saliendo del programa...')
        exit(1)


    else:
        logger.info('Success reading version of registry')


def get_output_directory(config, sampleID, assayID, determination, genes, cbio):

    db = cbio.tools.utils.hpc0_db.hpc0db_conn()

    try:
        db.insert_assay({"assayID": assayID})
    except:
        pass

    ID = db.get_id_sample_assay(sampleID, assayID, determination)

    if ID is None:
        ID = str(int(db.get_last_sample_id()) + 1)

    try:
        db.insert_sample_determination({
                                    'ID': ID,
                                    'sampleID': sampleID,
                                    'assayID': assayID,
                                    'determination': determination,
                                    'genes': genes
                                    })
    except:
        print('#[LOG]: Sample ' + str(sampleID) + ' already in database')

    db.close()

    output_directory = os.path.join(config['output_dir'], '_'.join([ID, sampleID, assayID]))
    return output_directory


def extract_information(ws, config, assayID, logger, cbio):
    """
    Function to extract list of genes, determination, index1, index2 and
    plasmids from the input excel file.
    These values start on row 10 and there are as many values as samples.
    """
    # Put it in extra file
    list_index = ['N701',
                  'N702',
                  'N703',
                  'N704',
                  'N705',
                  'N706',
                  'N707',
                  'N710',
                  'N711',
                  'N712',
                  'N714',
                  'N715',
                  'N716',
                  'N718',
                  'N719',
                  'N720',
                  'N721',
                  'N722',
                  'N723',
                  'N724',
                  'N726',
                  'N727',
                  'N728',
                  'N729',
                  'S502',
                  'S503',
                  'S505',
                  'S506',
                  'S507',
                  'S508',
                  'S510',
                  'S511',
                  'S513',
                  'S515',
                  'S516',
                  'S517',
                  'S518',
                  'S520',
                  'S521',
                  'S522']

    # Initiate empty dictionary to introduce values for all samples
    samples = {}

    # Initiate lists of indexes ir order to evaluate them
    index1_list = []
    index2_list = []

    # Iterate through the rows beginning in the one after the header
    for row in range (10, 1000):

        # If we are in the row 1000th it means that something went wrong
        if row > 1000:
            logger.error('Se ha llegado a la fila 1000 en el documento excel introducido')
            logger.error('Saliendo del programa...')
            exit(1)

        # Get sample + [genes] from the registry
        sample_genes = ws["{}{}".format("D", row)].value

        # If the value of the column "D" is not empty, parse all the information
        # in the row
        if sample_genes is not None:

            # Separate sample and genes in one step
            sample, genes = sample_genes.split(" ")

            # Add all the information of the sample to the dictionary
            samples[str(sample)] = {}
            samples[sample]['sampleID'] = str(sample)                              # Sample ID
            samples[sample]['genes'] = genes.split(';')                            # List of genes
            samples[sample]['stid'] = ws["{}{}".format("E", row)].value            # Plasmid ID
            samples[sample]['determination'] = ws["{}{}".format("O", row)].value   # Determination
            samples[sample]['outdir'] = get_output_directory(config, str(sample), assayID, samples[sample]['determination'], genes.split(';'), cbio)                     # Output Directory
            samples[sample]['assayID'] = assayID

            index1_value = ws["{}{}".format("M", row)].value                       # Index 1
            samples[sample]['index1'] = index1_value                               # Index 1
            index1_list.append(index1_value)                                       # Index 1

            index2_value = ws["{}{}".format("N", row)].value                       # Index 2
            samples[sample]['index2'] = index2_value                               # Index 2
            index2_list.append(index2_value)                                       # Index 2


        # If a row has an empty value in column "D", it means that we have
        # arrived to the end of the samples in the registry
        elif sample_genes is None:
            logger.info('Finished parsing excel file')

            # Break loop and return to main process
            break

    # Checks if index is in reference
    for elemento in index1_list:
        if elemento not in list_index:
            logger.error('Uno de los indices del PNT no estan en la referencia: ')
            logger.error(list_index)
            logger.error('Ha introducido:')
            logger.error(index_1)
            logger.error('Saliendo del programa...')
            exit(1)

    # Checks if index is in reference
    for elemento in index2_list:
        if elemento not in list_index:
            logger.error('Los indices del PNT no estan en la referencia: ')
            logger.error(list_index)
            logger.error('Saliendo del programa...')
            exit(1)

    logger.debug(samples)
    return(samples)


def excel_qc(excel_file_path, config, logger, cbio):
    """
    Function calling to col_names, pnt_version and extract_genes to check
    that everything is OK and extract all components of the returning dictionary

    This excel starts at line 9
    The wanted header column contains
        0	Muestra
        1	Genes
        2	Cód Plásmido
        3	Nº Amplicones
        4	Proporción (%)
        5	Qubit (ng/ul)
        6	Vf (µl)
        7	Cf (ng/uL)
        8	Vi (µl)
        9	Agua (ul)
        10	Index 1
        11	Index 2
        12	Determinación

    Columns to parse: 0, 1, 2, 10, 11, 12

    """

    if not os.path.exists(os.path.abspath(str(excel_file_path))):
        logger.error('Excel file not found in path -> ' + str(excel_file_path))
        logger.error('Exiting...')
        exit(1)

    wb = load_workbook(excel_file_path, read_only = True)

    assayID = excel_file_path.strip('.xlsx').split(' ')[-1].strip('E')

    ws = wb['Muestras']
    col_names(ws, logger)
    pnt_version(ws, logger)
    samples_dict = extract_information(ws, config, assayID, logger, cbio)

    return(samples_dict, assayID)

def parse_options(data):
        parser = argparse.ArgumentParser()
        args = parser.parse_args()
        # Do args checkings

        # Populate options with args atributes
        options={} # Add args atributes
        return options
